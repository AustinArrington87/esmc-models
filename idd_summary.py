import pandas as pd
import openpyxl
from openpyxl.styles import Font

# Load the dataset
file_path = 'Analytics_IDD_2023.csv'  # Update the path if necessary
data = pd.read_csv(file_path)

# Define the practice change categories
categories = ['Nutrient management', 'Normal operation', 'Tillage reduction', 'Cover cropping']

# Filter for rows with status "READY_TO_BE_VERIFIED"
data = data[data['model_dndc_s3s_submission_status__value'] == 'READY_TO_BE_VERIFIED']

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Function to process and add a category to the Excel workbook
def process_category(category):
    # Filter the data for the current category
    filtered_data = data[data['practice_change'] == category]

    # Group the data by program_region_crop and calculate both the number of fields and the summed acres
    grouped_data = filtered_data.groupby('program_region_crop').agg(
        Number_of_Fields=('id', 'nunique'),
        Summed_Acres=('acres', 'sum')
    ).reset_index()

    # Round the acres to two decimal places
    grouped_data['Summed_Acres'] = grouped_data['Summed_Acres'].round(2)

    # Create a new worksheet for the category
    worksheet = workbook.create_sheet(title=category)

    # Add header row
    worksheet.append(['Program Region Crop', 'Number of Fields', 'Summed Acres'])

    # Add data rows
    for index, row in grouped_data.iterrows():
        worksheet.append([
            row['program_region_crop'],
            row['Number_of_Fields'],
            f"{row['Summed_Acres']:.2f}"
        ])

    # Calculate totals
    total_fields = grouped_data['Number_of_Fields'].sum()
    total_acres = grouped_data['Summed_Acres'].sum()

    # Add a blank row
    worksheet.append([])

    # Add total row
    total_row = worksheet.max_row + 1
    worksheet.cell(row=total_row, column=1, value='TOTAL')
    worksheet.cell(row=total_row, column=2, value=total_fields)
    worksheet.cell(row=total_row, column=3, value=f"{total_acres:.2f}")

    # Apply bold font to the total row
    for cell in worksheet[total_row]:
        cell.font = Font(bold=True)

# Process each category and add it to the Excel workbook
for category in categories:
    process_category(category)

# Remove the default sheet created by openpyxl
workbook.remove(workbook['Sheet'])

# Save the Excel workbook
workbook.save('Agricultural_Practices_Summary.xlsx')
